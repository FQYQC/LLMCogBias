import csv
import openpyxl
import datetime
import json


def read_data(source_file):
    wb = openpyxl.load_workbook(source_file)
    ws = wb.active
    data = []
    for row in ws.iter_rows(
        min_row=0, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        data.append([cell.value for cell in row])
    return data


def main(source_file):

    data = read_data(source_file)
    title = data[0]
    t2c = {title[i]: i for i in range(len(title))}
    data = data[1:]

    # create effect_catagory list
    effect_catagory = []
    for row in data:
        if (
            row[t2c["effect_catagory"]] not in effect_catagory
            and row[t2c["text"]] is not None
        ):
            effect_catagory.append(row[t2c["effect_catagory"]])

    # create effect_name list and cata-map
    effect_name = []
    effect_map = {}
    for row in data:
        if row[t2c["effect_name"]] not in effect_name and row[t2c["text"]] is not None:
            effect_name.append(row[t2c["effect_name"]])
            effect_map[row[t2c["effect_name"]]] = row[t2c["effect_catagory"]]

    # create effect info json
    effect_info = []
    for effect in effect_name:
        effect_info.append({"name": effect, "catagory": effect_map[effect]})

    # create task
    tasks = []
    curr_id = 0
    sub_task_id = {effect: 0 for effect in effect_name}
    for row in data:
        if row[t2c["text"]] is None:
            continue
        task = {
            "id": curr_id,
            "text": row[t2c["text"]],
            "effect_name": row[t2c["effect_name"]],
            "effect_catagory": row[t2c["effect_catagory"]],
            "sub_task_id": sub_task_id[row[t2c["effect_name"]]],
            "answer_type": row[t2c["answer_type"]],
            "template": row[t2c["template"]],
            "num_variants": row[t2c["num_variants"]],
            "variants": [],
        }
        if task["num_variants"] is not None and task["num_variants"] > 0:
            task["variants"].append(row[t2c["var_value_1"]])
            for i in range(1, task["num_variants"]):
                task["variants"].append(row[t2c["var_value_1"] + i])
        sub_task_id[row[t2c["effect_name"]]] += 1
        curr_id += 1

        tasks.append(task)

    # create data json
    data_json = {
        "effect_catagory": effect_catagory,
        "effect_info": effect_info,
        "tasks": tasks,
    }

    with open(
        "datasets/data_{}.json".format(
            datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        ),
        "w",
    ) as f:
        json.dump(data_json, f, indent=4)


if __name__ == "__main__":
    main("/mnt/d/QYQC/PKU/GradPaper/experiments/data.xlsx")
